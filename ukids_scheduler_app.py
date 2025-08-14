# ukids_scheduler_app.py
# Robust CSV loader + slot-per-row scheduler
# - Accepts ANY CSV filenames for positions & availability
# - Tries multiple encodings (utf-8, utf-8-sig, cp1252, iso-8859-1)
# - Tries sniffed separators (engine='python', sep=None), then ',', ';', '\t', '|'
# - Each slot is its own row (leaders included)
# - Ignores priorities except 0 (0 = not eligible)
# - Max 2 assignments per person across the whole month
# - People with <2 "Yes" dates are shown on a separate sheet **for reference only** (NOT excluded)
# - Output sheet is named like "September 2025"
# - Columns auto-fit widths in Excel

import io
import re
import base64
from collections import defaultdict, Counter
from datetime import datetime, date

import numpy as np
import pandas as pd
import streamlit as st
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

st.set_page_config(page_title="uKids Scheduler", layout="wide")
st.title("uKids Scheduler")

st.markdown(
    """
    <style>
      .stApp { background: #000; color: #fff; }
      .stButton>button, .stDownloadButton>button { background:#444; color:#fff; }
      .stDataFrame { background:#111; }
      .stAlert { color:#111; }
    </style>
    """,
    unsafe_allow_html=True,
)

# Optional logo (ignore if missing)
for logo_name in ["image(1).png", "image.png", "logo.png"]:
    try:
        with open(logo_name, "rb") as img_file:
            encoded = base64.b64encode(img_file.read()).decode()
            st.markdown(
                f"<div style='text-align:center'><img src='data:image/png;base64,{encoded}' width='520'></div>",
                unsafe_allow_html=True,
            )
            break
    except Exception:
        pass

# ──────────────────────────────────────────────────────────────────────────────
# Helpers
# ──────────────────────────────────────────────────────────────────────────────
MONTH_ALIASES = {
    "jan": 1, "january": 1,
    "feb": 2, "february": 2,
    "mar": 3, "march": 3,
    "apr": 4, "april": 4,
    "may": 5,
    "jun": 6, "june": 6,
    "jul": 7, "july": 7,
    "aug": 8, "august": 8,
    "sep": 9, "sept": 9, "september": 9,
    "oct": 10, "october": 10,
    "nov": 11, "november": 11,
    "dec": 12, "december": 12,
}
YES_SET = {"yes", "y", "true", "available"}

def read_csv_robust(uploaded_file, label_for_error):
    """
    Read a Streamlit UploadedFile into a DataFrame, trying multiple encodings and separators.
    Raises a clear Streamlit error if everything fails.
    """
    raw = uploaded_file.getvalue()
    encodings = ["utf-8", "utf-8-sig", "cp1252", "iso-8859-1"]
    seps = [None, ",", ";", "\t", "|"]

    last_err = None
    for enc in encodings:
        for sep in seps:
            try:
                df = pd.read_csv(io.BytesIO(raw), encoding=enc, engine="python", sep=sep)
                if df.shape[1] == 0:
                    raise ValueError("Parsed 0 columns.")
                return df
            except Exception as e:
                last_err = f"{type(e).__name__}: {e}"
                continue
    st.error(
        f"Could not read {label_for_error} CSV. Last error: {last_err}. "
        "Try re-exporting as CSV (UTF-8) or remove unusual characters in headers."
    )
    st.stop()

def detect_name_column(df: pd.DataFrame, fallback_first: bool = True) -> str:
    candidates = [
        "What is your name AND surname?",
        "What is your name and surname?",
        "Name",
        "Full name",
        "Full names",
    ]
    cols_l = {str(c).strip().lower(): c for c in df.columns}
    for c in candidates:
        key = c.strip().lower()
        if key in cols_l:
            return cols_l[key]
    for c in df.columns:
        if isinstance(c, str) and "name" in c.lower():
            return c
    if fallback_first:
        return df.columns[0]
    raise ValueError("Could not detect a 'name' column.")

def is_priority_col(series: pd.Series) -> bool:
    vals = pd.to_numeric(series, errors="coerce").dropna()
    if len(vals) == 0:
        return False
    return (vals.min() >= 0) and (vals.max() <= 5)

def normalize(s: str) -> str:
    return re.sub(r"[^a-z0-9]+", " ", str(s).lower()).strip()

def parse_month_and_dates_from_headers(responses_df: pd.DataFrame):
    avail_cols = [c for c in responses_df.columns if isinstance(c, str) and c.strip().lower().startswith("are you available")]
    if not avail_cols:
        # also accept headers that contain a day number and a month name
        avail_cols = [
            c for c in responses_df.columns
            if isinstance(c, str)
            and re.search(r"\b\d{1,2}\b", c.lower())
            and any(m in c.lower() for m in MONTH_ALIASES)
        ]
    if not avail_cols:
        raise ValueError("No availability columns found. Expect headings like 'Are you available 7 September?'")

    info = []
    for c in avail_cols:
        low = c.lower()
        mname = None
        for alias in MONTH_ALIASES:
            if alias in low:
                mname = alias
                break
        day_m = re.search(r"\b(\d{1,2})\b", low)
        if mname and day_m:
            info.append((c, MONTH_ALIASES[mname], int(day_m.group(1))))
    if not info:
        raise ValueError("Could not parse day/month from availability headers.")

    months = {m for _, m, _ in info}
    if len(months) > 1:
        raise ValueError(f"Multiple months detected in availability headers: {sorted(months)}. Upload one month at a time.")
    month = months.pop()

    if "Timestamp" in responses_df.columns:
        years = pd.to_datetime(responses_df["Timestamp"], errors="coerce").dt.year.dropna().astype(int)
        year = int(years.mode().iloc[0]) if not years.empty else date.today().year
    else:
        year = date.today().year

    date_map = {c: pd.Timestamp(datetime(year, month, d)).normalize() for c, _, d in info}
    service_dates = sorted(set(date_map.values()))
    sheet_name = f"{pd.Timestamp(year=year, month=month, day=1):%B %Y}"
    return year, month, date_map, service_dates, sheet_name

def build_long_df(people_df: pd.DataFrame, name_col: str, role_cols):
    records = []
    for _, r in people_df.iterrows():
        person = str(r[name_col]).strip()
        if not person or person.lower() == "nan":
            continue
        for role in role_cols:
            pr = pd.to_numeric(r[role], errors="coerce")
            if pd.isna(pr):
                continue
            pr = int(round(pr))
            if pr >= 1:  # 0 means not an option
                records.append({"person": person, "role": role, "priority": pr})
    return pd.DataFrame(records)

def parse_availability(responses_df: pd.DataFrame, name_col_resp: str, date_map):
    availability = {}
    yes_counts = Counter()
    for _, row in responses_df.iterrows():
        nm = str(row.get(name_col_resp, "")).strip()
        if not nm or nm.lower() == "nan":
            continue
        availability.setdefault(nm, {})
        for col, dt in date_map.items():
            ans = str(row.get(col, "")).strip().lower()
            is_yes = ans in YES_SET
            availability[nm][dt] = is_yes
            if is_yes:
                yes_counts[nm] += 1
    few_yes = sorted([n for n, c in yes_counts.items() if c < 2])
    service_dates = sorted(set(date_map.values()))
    return availability, service_dates, few_yes

# ──────────────────────────────────────────────────────────────────────────────
# Slot plan (leaders included) — every slot becomes its own row
# ──────────────────────────────────────────────────────────────────────────────
def build_slot_plan():
    return {
        # Age 1
        "Age 1 leader": 1,
        "Age 1 classroom": 5,
        "Age 1 nappies": 1,
        "Age 1 bags girls": 1,
        "Age 1 bags boys": 1,
        # Age 2
        "Age 2 leader": 1,
        "Age 2 classroom": 4,
        "Age 2 nappies": 1,
        "Age 2 bags girls": 1,
        "Age 2 bags boys": 1,
        # Age 3
        "Age 3 leader": 1,
        "Age 3 classroom": 4,
        "Age 3 bags": 1,
        # Age 4
        "Age 4 leader": 1,
        "Age 4 classroom": 4,
        # Age 5
        "Age 5 leader": 1,
        "Age 5 classroom": 3,
        # Age 6
        "Age 6 leader": 1,
        "Age 6 classroom": 3,
        # Age 7
        "Age 7 leader": 1,
        "Age 7 classroom": 2,
        # Age 8
        "Age 8 leader": 1,
        "Age 8 classroom": 2,
        # Age 9 (two separate one-person rows)
        "Age 9 leader": 1,
        "Age 9 classroom A": 1,
        "Age 9 classroom B": 1,
        # Age 10
        "Age 10 leader": 1,
        "Age 10 classroom": 1,
        # Age 11
        "Age 11 leader": 1,
        "Age 11 classroom": 1,
        # Special Needs
        "Special needs leader": 1,
        "Special needs classroom": 2,
    }

def expand_roles_to_slots(slot_plan):
    slot_rows = []
    slot_index = {}
    for role, n in slot_plan.items():
        if n <= 0:
            continue
        if n == 1:
            lab = role
            slot_rows.append(lab)
            slot_index[lab] = role
        else:
            for i in range(1, n + 1):
                lab = f"{role} #{i}"
                slot_rows.append(lab)
                slot_index[lab] = role
    return slot_rows, slot_index

def build_eligibility(long_df: pd.DataFrame):
    elig = defaultdict(set)
    for _, r in long_df.iterrows():
        elig[str(r["person"]).strip()].add(str(r["role"]).strip())
    return elig

def schedule_by_slots(long_df, availability, service_dates, max_assignments_per_person=2):
    slot_plan = build_slot_plan()
    slot_rows, slot_to_role = expand_roles_to_slots(slot_plan)
    eligibility = build_eligibility(long_df)

    # Only schedule people that exist in both sources
    people = sorted(set(eligibility.keys()) & set(availability.keys()))

    grid = {(slot, d): "" for slot in slot_rows for d in service_dates}
    assign_count = defaultdict(int)

    for d in service_dates:
        assigned_today = set()
        for slot_row in slot_rows:
            base_role = slot_to_role[slot_row]

            cands = []
            for p in people:
                if assign_count[p] >= max_assignments_per_person:
                    continue
                if p in assigned_today:
                    continue
                if not availability.get(p, {}).get(d, False):
                    continue
                # exact or normalized match
                elig_roles = eligibility.get(p, set())
                ok = False
                if base_role in elig_roles:
                    ok = True
                else:
                    nb = normalize(base_role)
                    for er in elig_roles:
                        if normalize(er) == nb:
                            ok = True
                            break
                if ok:
                    cands.append(p)

            if cands:
                # Prioritize least-used volunteers first
                cands.sort(key=lambda name: assign_count[name])
                chosen = cands[0]
                grid[(slot_row, d)] = chosen
                assign_count[chosen] += 1
                assigned_today.add(chosen)
            # else: leave empty if no eligible/available person

    cols = [d.strftime("%Y-%m-%d") for d in service_dates]
    schedule_df = pd.DataFrame(index=slot_rows, columns=cols)
    for (slot_row, d), name in grid.items():
        schedule_df.loc[slot_row, d.strftime("%Y-%m-%d")] = name
    schedule_df = schedule_df.fillna("")
    return schedule_df, dict(assign_count)

def excel_autofit(ws):
    for col_idx, column_cells in enumerate(
        ws.iter_cols(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column), start=1
    ):
        max_len = 0
        for cell in column_cells:
            val = "" if cell.value is None else str(cell.value)
            max_len = max(max_len, len(val))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max(12, max_len + 2), 80)

# ──────────────────────────────────────────────────────────────────────────────
# UI
# ──────────────────────────────────────────────────────────────────────────────
st.subheader("1) Upload files (CSV — any filename)")

c1, c2 = st.columns(2)
with c1:
    positions_file = st.file_uploader("Serving positions (CSV)", type=["csv"], key="positions_csv_any")
with c2:
    responses_file = st.file_uploader("Availability responses (CSV)", type=["csv"], key="responses_csv_any")

st.caption("• Positions CSV: first column = volunteer names; other columns are role headings with values 0‑5 (0 = not eligible).")
st.caption("• Responses CSV: includes a name column (e.g., 'What is your name AND surname?') and columns like 'Are you available 7 September?'.")

if st.button("Generate Schedule", type="primary"):
    if not positions_file or not responses_file:
        st.error("Please upload both CSV files.")
        st.stop()

    positions_df = read_csv_robust(positions_file, "positions")
    responses_df = read_csv_robust(responses_file, "responses")

    # Detect name columns
    try:
        name_col_positions = detect_name_column(positions_df, fallback_first=True)
    except Exception as e:
        st.error(f"Could not detect a name column in positions CSV: {e}")
        st.stop()

    try:
        name_col_responses = detect_name_column(responses_df, fallback_first=False)
    except Exception as e:
        st.error(f"Could not detect a name column in responses CSV: {e}")
        st.stop()

    # Ensure name columns are strings
    positions_df[name_col_positions] = positions_df[name_col_positions].astype(str)
    responses_df[name_col_responses] = responses_df[name_col_responses].astype(str)

    # Role columns
    role_cols = [c for c in positions_df.columns if c != name_col_positions and is_priority_col(positions_df[c])]
    if not role_cols:
        st.error("No role columns with priorities (0..5) detected in the positions CSV.")
        st.stop()

    # Build eligibility
    long_df = build_long_df(positions_df, name_col_positions, role_cols)
    if long_df.empty:
        st.error("No eligible assignments found (all priorities are 0 or missing).")
        st.stop()

    # Parse dates
    try:
        year, month, date_map, service_dates, sheet_name = parse_month_and_dates_from_headers(responses_df)
    except Exception as e:
        st.error(f"Could not parse month & dates from responses: {e}")
        st.stop()

    # Availability & <2-Yes list (INFO ONLY — not excluded)
    availability, service_dates, few_yes_list = parse_availability(responses_df, name_col_responses, date_map)

    # Build schedule (no exclusions for one-Yes people)
    schedule_df, assign_count = schedule_by_slots(
        long_df, availability, service_dates, max_assignments_per_person=2
    )

    # Stats
    total_slots = schedule_df.size
    filled_slots = int((schedule_df != "").sum().sum())
    fill_rate = (filled_slots / total_slots) if total_slots else 0.0
    unfilled = total_slots - filled_slots
    per_person = pd.Series(assign_count, name="Assignments").sort_values(ascending=False).reset_index().rename(columns={"index":"Person"})

    st.success(f"Schedule generated for **{sheet_name}**")
    st.write(f"Filled slots: **{filled_slots} / {total_slots}**  (Fill rate: **{fill_rate:.1%}**)  •  Unfilled: **{unfilled}**")

    st.subheader("Schedule (each slot is its own row)")
    st.dataframe(schedule_df, use_container_width=True)

    st.subheader("Assignment Summary")
    st.dataframe(per_person, use_container_width=True)

    st.subheader("People with < 2 'Yes' dates (for reference only)")
    few_yes_df = pd.DataFrame({"Person": few_yes_list})
    st.dataframe(few_yes_df, use_container_width=True)

    # Excel output
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name

    header = ["Position / Slot"] + [d.strftime("%Y-%m-%d") for d in service_dates]
    ws.append(header)
    for idx, row_name in enumerate(schedule_df.index, start=2):
        row_vals = [row_name] + [schedule_df.iloc[idx-2, j] for j in range(len(service_dates))]
        ws.append(row_vals)
    excel_autofit(ws)

    if few_yes_list:
        ws2 = wb.create_sheet("Fewer than 2 Yes (info)")
        ws2.append(["Person"])
        for p in few_yes_list:
            ws2.append([p])
        excel_autofit(ws2)

    ws3 = wb.create_sheet("Assignment Summary")
    ws3.append(["Person", "Assignments"])
    for _, r in per_person.iterrows():
        ws3.append([r["Person"], int(r["Assignments"])])
    excel_autofit(ws3)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    st.download_button(
        "Download Excel (.xlsx)",
        data=buf,
        file_name=f"uKids_schedule_{sheet_name.replace(' ','_')}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
else:
    st.info("Upload the two CSV files (any names), then click **Generate Schedule**.")
